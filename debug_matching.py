import tkinter as tk  
from tkinter import filedialog, messagebox  
from openpyxl import load_workbook  
from dbf import Table, READ_WRITE  
  
# --- Helper function to clean numeric values ---  
def clean_value(val):  
    if val is None:  
        return ''  
    if isinstance(val, (int, float)):  
        return f"{val:.1f}"  
    return val  
  
def debug_process_files():  
    excel_path = r"d:\Python\ECLASSRECORD2DBF WEB\testfiles\2506B.xlsm"  
    dbf_path = r"d:\Python\ECLASSRECORD2DBF WEB\testfiles\DSO_20243_2506B_BACC104_565.DBF" 
  
    if not excel_path or not dbf_path:  
        print("Please select both files.")  
        return  
  
    try:  
        # Load Excel  
        wb = load_workbook(excel_path, data_only=True)  
        print(f"Excel sheets: {wb.sheetnames}")  
  
        ws = wb["FFG"]

        # Find the columns with "EG" and "REMARKS" in row 7
        col_grade_idx = None
        col_remark_idx = None

        for col_idx in range(1, ws.max_column + 1):  # Iterate through all columns
            cell_value = ws.cell(row=7, column=col_idx).value
            if cell_value and str(cell_value).strip().upper() == "EG":
                col_grade_idx = col_idx
            elif cell_value and str(cell_value).strip().upper() == "REMARKS":
                col_remark_idx = col_idx

        if col_grade_idx is None:
            raise ValueError("Column with 'EG' header not found in row 7")
        if col_remark_idx is None:
            raise ValueError("Column with 'REMARKS' header not found in row 7")

        # Convert column indices to letters for easier reference
        def col_to_letter(col_num):
            result = ""
            while col_num > 0:
                col_num -= 1
                result = chr(col_num % 26 + ord('A')) + result
                col_num //= 26
            return result

        col_grade_letter = col_to_letter(col_grade_idx)
        col_remark_letter = col_to_letter(col_remark_idx)

        print(f"Using worksheet: FFG")
        print(f"Found 'EG' in column {col_grade_letter} and 'REMARKS' in column {col_remark_letter}")

        row = 11
        excel_data = {}
        print("Reading Excel data from column C ^(ID^) and dynamic grade/remark columns:")

        while True:
            cell_val = ws[f'C{row}'].value
            if cell_val is None:
                print(f"Stopping at row {row} - cell C{row} is empty")
                break

            print(f"Row {row}, C{row} = {cell_val}, type: {type(cell_val)}")

            try:
                # This is the current approach in your code - converting to int
                id_str = str(int(cell_val)).strip()
                print(f"  Converted ID: {id_str}")
            except Exception as e:
                print(f"  Skipping row {row} - conversion failed: {e}")
                row += 1
                continue

            col_grade = ws[f'{col_grade_letter}{row}'].value
            col_remark = ws[f'{col_remark_letter}{row}'].value
            print(f"  {col_grade_letter}{row} = {col_grade}, {col_remark_letter}{row} = {col_remark}")
            excel_data[id_str] = (col_grade, col_remark)
            row += 1

            # Just process first 5 non-empty rows for debugging
            if len(excel_data) >= 5:
                break
  
        print(f"Excel data collected: {excel_data}")  
        wb.close()  
  
        # Open DBF for read/write  
        table = Table(dbf_path)  
        table.open(mode=READ_WRITE) 
  
        print(f"DBF field names: {table.field_names}")  
        id_index = 5  # 5th column, 0-based index  
        print(f"Checking field at index {id_index}: {table.field_names[id_index] if id_index < len(table.field_names) else 'FIELD INDEX OUT OF RANGE'}")  
  
        target_col3 = 2  # 3rd column (write H)  
        target_col4 = 3  # 4th column (write I)  
  
        print(f"Target columns: {table.field_names[target_col3]} (index {target_col3}) and {table.field_names[target_col4]} (index {target_col4})")  
  
        matched = 0  
        processed = 0  
  
        for i, record in enumerate(table):  
            processed += 1  
            dbf_id = str(record[id_index]).strip()  
            print(f"DBF record {i+1}, ID at field {id_index}: '{dbf_id}' (type: {type(record[id_index])})")  
  
            if dbf_id in excel_data:  
                h_val, i_val = excel_data[dbf_id]  
                print(f"  MATCH found! DBF ID '{dbf_id}' matches Excel data: H={h_val}, I={i_val}")  
                matched += 1  
                # Don't actually modify for this debug
                if matched >= 5:  # Just check first few matches
                    break
            else:  
                print(f"  No match for DBF ID '{dbf_id}' in Excel data")  
  
            if processed >= 10:  # Just check first 10 records
                break
  
        table.close()  
        print(f"Total DBF records processed: {processed}, matches found: {matched}")  
  
    except Exception as e:  
        print(f"Error: {str(e)}")  
        import traceback  
        traceback.print_exc()  
  
if __name__ == "__main__":  
    debug_process_files() 
