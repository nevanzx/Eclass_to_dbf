import streamlit as st
from openpyxl import load_workbook
from dbf import Table, READ_WRITE
import tempfile
import os
import pandas as pd
import io
import base64
from reports import show_word_report_ui, convert_word_to_pdf, generate_word_report
from config import extract_jle_data

# --- Helper function to clean numeric values ---
def clean_value(val):
    if val is None:
        return ''
    if isinstance(val, (int, float)):
        return f"{val:.1f}"
    return val

def process_files(excel_file, dbf_file, original_dbf_filename):
    """Process the Excel and DBF files based on the original logic"""
    excel_path = None
    dbf_path = None

    try:
        # Save uploaded files to temporary locations
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_excel:
            tmp_excel.write(excel_file.getvalue())
            excel_path = tmp_excel.name

        with tempfile.NamedTemporaryFile(delete=False, suffix='.dbf') as tmp_dbf:
            tmp_dbf.write(dbf_file.getvalue())
            dbf_path = tmp_dbf.name

        # Load Excel - using context manager to ensure it's properly closed
        wb = load_workbook(excel_path, data_only=True)
        try:
            # Check if the "FFG" worksheet exists
            if "FFG" not in wb.sheetnames:
                available_sheets = ", ".join(wb.sheetnames)
                raise ValueError(f"Worksheet 'FFG' not found. Available sheets: {available_sheets}")

            ws = wb["FFG"]

            # Find the columns with "EG" and "REMARKS" in row 7
            col_grade_idx = None
            col_remark_idx = None

            for col_idx in range(1, ws.max_column + 1):  # Iterate through all columns
                cell_value = ws.cell(row=7, column=col_idx).value
                if cell_value and str(cell_value).strip().upper() == "EG":
                    col_grade_idx = col_idx
                elif cell_value and str(cell_value).strip().upper() == "REMARKS":
                    col_remark_idx = col_idx

            if col_grade_idx is None:
                raise ValueError("Column with 'EG' header not found in row 7")
            if col_remark_idx is None:
                raise ValueError("Column with 'REMARKS' header not found in row 7")

            # Convert column indices to letters for easier reference
            def col_to_letter(col_num):
                result = ""
                while col_num > 0:
                    col_num -= 1
                    result = chr(col_num % 26 + ord('A')) + result
                    col_num //= 26
                return result

            col_grade_letter = col_to_letter(col_grade_idx)
            col_remark_letter = col_to_letter(col_remark_idx)

            row = 11
            excel_data = {}

            while True:
                cell_val = ws[f'C{row}'].value
                if cell_val is None:
                    break

                try:
                    id_str = str(int(cell_val)).strip()
                except (ValueError, TypeError):
                    # Skip rows where C column doesn't contain a valid integer
                    row += 1
                    continue

                col_grade = ws[f'{col_grade_letter}{row}'].value
                col_remark = ws[f'{col_remark_letter}{row}'].value
                excel_data[id_str] = (col_grade, col_remark)
                row += 1
        finally:
            wb.close()  # Ensure workbook is closed even if an exception occurs

        # Open DBF for read/write
        table = Table(dbf_path)
        try:
            table.open(mode=READ_WRITE)

            id_index = 5  # 5th column, 0-based index
            target_col3 = 2  # 3rd column (write grade)
            target_col4 = 3  # 4th column (write remark)

            matched = 0
            for record in table:
                dbf_id = str(record[id_index]).strip()
                if dbf_id in excel_data:
                    grade_val, remark_val = excel_data[dbf_id]
                    with record:  # <- Required for safe writing
                        if grade_val is not None:
                            record[target_col3] = clean_value(grade_val)
                        if remark_val is not None:
                            record[target_col4] = clean_value(remark_val)
                    matched += 1
        finally:
            table.close()  # Ensure table is closed even if an exception occurs

        # Copy the updated DBF file back to bytes
        with open(dbf_path, 'rb') as f:
            updated_dbf_bytes = f.read()

        return updated_dbf_bytes, matched

    finally:
        # Ensure temporary files are cleaned up in all cases
        if excel_path and os.path.exists(excel_path):
            try:
                os.remove(excel_path)
            except Exception:
                # If we can't remove the file, it might be locked by another process
                pass

        if dbf_path and os.path.exists(dbf_path):
            try:
                os.remove(dbf_path)
            except Exception:
                # If we can't remove the file, it might be locked by another process
                pass




def process_files_with_jle(jle_data, excel_file, dbf_file, original_dbf_filename):
    """
    Process the Excel and DBF files with additional JLE data.
    This function extends the original process_files function to incorporate JLE data.
    """
    excel_path = None
    dbf_path = None

    try:
        # Save uploaded files to temporary locations
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_excel:
            tmp_excel.write(excel_file.getvalue())
            excel_path = tmp_excel.name

        with tempfile.NamedTemporaryFile(delete=False, suffix='.dbf') as tmp_dbf:
            tmp_dbf.write(dbf_file.getvalue())
            dbf_path = tmp_dbf.name

        # Load Excel - using context manager to ensure it's properly closed
        wb = load_workbook(excel_path, data_only=True)
        try:
            # Check if the "FFG" worksheet exists
            if "FFG" not in wb.sheetnames:
                available_sheets = ", ".join(wb.sheetnames)
                raise ValueError(f"Worksheet 'FFG' not found. Available sheets: {available_sheets}")

            ws = wb["FFG"]

            # Find the columns with "EG" and "REMARKS" in row 7
            col_grade_idx = None
            col_remark_idx = None

            for col_idx in range(1, ws.max_column + 1):  # Iterate through all columns
                cell_value = ws.cell(row=7, column=col_idx).value
                if cell_value and str(cell_value).strip().upper() == "EG":
                    col_grade_idx = col_idx
                elif cell_value and str(cell_value).strip().upper() == "REMARKS":
                    col_remark_idx = col_idx

            if col_grade_idx is None:
                raise ValueError("Column with 'EG' header not found in row 7")
            if col_remark_idx is None:
                raise ValueError("Column with 'REMARKS' header not found in row 7")

            # Convert column indices to letters for easier reference
            def col_to_letter(col_num):
                result = ""
                while col_num > 0:
                    col_num -= 1
                    result = chr(col_num % 26 + ord('A')) + result
                    col_num //= 26
                return result

            col_grade_letter = col_to_letter(col_grade_idx)
            col_remark_letter = col_to_letter(col_remark_idx)

            row = 11
            excel_data = {}

            while True:
                cell_val = ws[f'C{row}'].value
                if cell_val is None:
                    break

                try:
                    id_str = str(int(cell_val)).strip()
                except (ValueError, TypeError):
                    # Skip rows where C column doesn't contain a valid integer
                    row += 1
                    continue

                col_grade = ws[f'{col_grade_letter}{row}'].value
                col_remark = ws[f'{col_remark_letter}{row}'].value
                excel_data[id_str] = (col_grade, col_remark)
                row += 1
        finally:
            wb.close()  # Ensure workbook is closed even if an exception occurs

        # Open DBF for read/write
        table = Table(dbf_path)
        try:
            table.open(mode=READ_WRITE)

            id_index = 5  # 5th column, 0-based index
            target_col3 = 2  # 3rd column (write grade)
            target_col4 = 3  # 4th column (write remark)

            matched = 0
            for record in table:
                dbf_id = str(record[id_index]).strip()
                if dbf_id in excel_data:
                    grade_val, remark_val = excel_data[dbf_id]
                    with record:  # <- Required for safe writing
                        if grade_val is not None:
                            record[target_col3] = clean_value(grade_val)
                        if remark_val is not None:
                            record[target_col4] = clean_value(remark_val)
                    matched += 1
        finally:
            table.close()  # Ensure table is closed even if an exception occurs

        # Copy the updated DBF file back to bytes
        with open(dbf_path, 'rb') as f:
            updated_dbf_bytes = f.read()

        return updated_dbf_bytes, matched

    finally:
        # Ensure temporary files are cleaned up in all cases
        if excel_path and os.path.exists(excel_path):
            try:
                os.remove(excel_path)
            except Exception:
                # If we can't remove the file, it might be locked by another process
                pass

        if dbf_path and os.path.exists(dbf_path):
            try:
                os.remove(dbf_path)
            except Exception:
                # If we can't remove the file, it might be locked by another process
                pass


def read_dbf_to_dataframe(dbf_bytes):
    """Convert DBF bytes to a pandas DataFrame for display"""
    # Create a temporary file to work with the DBF data
    with tempfile.NamedTemporaryFile(delete=False, suffix='.dbf') as tmp_dbf:
        tmp_dbf.write(dbf_bytes)
        temp_dbf_path = tmp_dbf.name

    try:
        # Open the DBF file
        table = Table(temp_dbf_path)
        table.open(mode=READ_WRITE)

        # Get field names
        field_names = table.field_names

        # Create a list to store records
        records = []
        for record in table:
            # Convert record to dictionary and add to records list
            # Handle the conversion carefully to avoid field access issues
            record_dict = {}
            for field_name in field_names:
                try:
                    record_dict[field_name] = record[field_name]
                except:
                    # If there's an issue accessing the field, set to None
                    record_dict[field_name] = None
            records.append(record_dict)

        # Close the table
        table.close()

        # Create a DataFrame from the records
        if records:
            df = pd.DataFrame(records, columns=field_names)
        else:
            # If no records, create empty dataframe with proper columns
            df = pd.DataFrame(columns=field_names)
        return df

    finally:
        # Clean up the temporary file
        if os.path.exists(temp_dbf_path):
            try:
                os.remove(temp_dbf_path)
            except Exception:
                pass










def main():
    st.set_page_config(
        page_title="E-Class DBF Updater",
        layout="wide",
        initial_sidebar_state="collapsed"
    )
    st.title("E-Class DBF Updater")
    st.markdown("Upload your RAR file containing JLE and DBF files, and Excel file separately to update grades automatically.")

    # Add instructions
    with st.expander("How to use this tool", expanded=True):
        st.markdown("""
        **Step-by-step instructions:**
        1. Upload your **JLE file and multiple DBF files** together (select multiple files)
        2. Select the appropriate **DBF file** from the dropdown
        3. Upload your **Excel file** (E-Class record with grades and remarks)
        4. Click **Update DBF** to process and update the grade sheet
        """)

        st.info("Note: Upload one JLE file and multiple DBF files together, then select which DBF file to process.")

    # Card-based layout for Steps 1 and 2
    col1, col2 = st.columns(2)

    with col1:
        with st.container(border=True):
            st.subheader("📁 Step 1: Upload JLE and DBF Files")

            uploaded_files = st.file_uploader(
                "Select JLE and multiple DBF files",
                type=['jle', 'dbf'],
                accept_multiple_files=True,
                key='multi_files'
            )

            # Process uploaded files to separate JLE and DBF files
            jle_file = None
            dbf_files = {}
            selected_dbf_file = None
            selected_dbf_name = None

            if uploaded_files:
                # Separate JLE and DBF files
                jle_candidates = []
                dbf_candidates = []

                for file in uploaded_files:
                    if file.name.lower().endswith('.jle'):
                        jle_candidates.append(file)
                    elif file.name.lower().endswith('.dbf'):
                        dbf_candidates.append(file)

                # Validate that we have exactly one JLE file
                if len(jle_candidates) == 0:
                    st.error("❌ No JLE file found! Please upload one JLE file along with DBF files.")
                elif len(jle_candidates) > 1:
                    st.error(f"❌ Multiple JLE files found ({len(jle_candidates)}). Please upload only one JLE file.")
                else:
                    jle_file = jle_candidates[0]
                    st.success(f"✅ JLE file: {jle_file.name}")

                    # Process DBF files
                    if len(dbf_candidates) == 0:
                        st.error("❌ No DBF files found! Please upload at least one DBF file along with the JLE file.")
                    else:
                        st.success(f"✅ Found {len(dbf_candidates)} DBF file(s)")

                        # Store JLE file and DBF candidates in session state for Step 3
                        st.session_state.jle_file = jle_file
                        st.session_state.dbf_candidates = dbf_candidates

    with col2:
        with st.container(border=True):
            st.subheader("📊 Step 2: Upload Excel File")

            excel_file = st.file_uploader("E-Class Record (Excel)", type=['xlsx', 'xlsm', 'xls'], key='excel')

            if excel_file:
                st.success(f"✅ Excel file: {excel_file.name}")
                # Store Excel file content in session state (not the file object itself)
                st.session_state.excel_content = excel_file.getvalue()
                st.session_state.excel_filename = excel_file.name
            else:
                st.info("ℹ️ Please upload your Excel file with grades and remarks")

    # Check if JLE and DBF files match
    if selected_dbf_name and jle_file:
        try:
            # Extract parts from selected DBF filename to check for match
            import re
            parts = selected_dbf_name.replace('.DBF', '').replace('.dbf', '').split('_')
            if len(parts) >= 4:
                year_semester = parts[1]  # YYYYX
                subj_num = parts[2]       # SUBJNUM
                subj_code = parts[3]      # SUBJCODE

                # Check against JLE data
                jle_data = extract_jle_data(io.BytesIO(jle_file.getvalue()))

                if 'course_data' in jle_data and jle_data['course_data'] is not None and not jle_data['course_data'].empty:
                    jle_df = jle_data['course_data']

                    match_found = False
                    for _, jle_record in jle_df.iterrows():
                        jle_academic_year = jle_record.get('Academic Year', '')  # Full format like "2024-2025"
                        if jle_academic_year:
                            jle_year = jle_academic_year.split('-')[0]  # Get first part like "2024"
                            from reports import get_semester_digit
                            jle_year_semester = f"{jle_year}{get_semester_digit(jle_record.get('Semester', ''))}"  # Full format like "20243"
                            jle_subj_num = jle_record.get('Subject Num', '')
                            jle_subj_code = jle_record.get('Subject Code', '')

                            # Check if the components match
                            if (year_semester == jle_year_semester and
                                subj_num == jle_subj_num and
                                subj_code == jle_subj_code):
                                match_found = True
                                st.success(f"✓ Match: {jle_subj_code}")
                                break

                    if not match_found:
                        st.error("✗ No match")
                else:
                    st.text("?")
            else:
                st.text("?")

        except Exception as e:
            st.text(f"Error: {str(e)}")

    # Step 3: Select DBF file and update
    with st.container(border=True):
        st.subheader("🔄 Step 3: Select DBF File and Update")

        # Check if we have DBF candidates stored in session state
        if 'dbf_candidates' in st.session_state and st.session_state.dbf_candidates:
            # Create a dropdown to select the DBF file
            dbf_options = [file.name for file in st.session_state.dbf_candidates]
            selected_dbf_name = st.selectbox("Select DBF file to process:", dbf_options)

            if selected_dbf_name:
                # Find the selected DBF file
                for file in st.session_state.dbf_candidates:
                    if file.name == selected_dbf_name:
                        selected_dbf_file = file
                        break

                if selected_dbf_file:
                    st.success(f"✅ DBF file selected: {selected_dbf_name}")

                    # Store in session state
                    st.session_state.selected_dbf_file = selected_dbf_file
                    st.session_state.selected_dbf_name = selected_dbf_name

            # Update DBF button in this card
            if st.button("📊 Update DBF", type="primary", key="update_dbf_step3"):
                # Check if all required files are provided
                if ('selected_dbf_file' not in st.session_state or
                    'excel_content' not in st.session_state or
                    'jle_file' not in st.session_state):
                    st.warning("Please upload JLE and DBF files, select a DBF file, and upload an Excel file.")
                else:
                    try:
                        with st.spinner('Processing files...'):
                            # Get the JLE file from session state (in the multiple file approach)
                            if 'jle_file' in st.session_state and st.session_state.jle_file:
                                jle_content = st.session_state.jle_file.getvalue()
                                jle_filename = st.session_state.jle_file.name
                            else:
                                st.error("JLE file not found. Please re-upload your files.")
                                st.stop()

                            # Create temporary files for the JLE and selected DBF content
                            with tempfile.NamedTemporaryFile(delete=False, suffix='.jle') as tmp_jle:
                                tmp_jle.write(jle_content)
                                jle_path = tmp_jle.name

                            with tempfile.NamedTemporaryFile(delete=False, suffix='.dbf') as tmp_dbf:
                                # Write the content of the selected DBF file
                                dbf_content = st.session_state.selected_dbf_file.getvalue()
                                tmp_dbf.write(dbf_content)
                                dbf_path = tmp_dbf.name

                            try:
                                # Extract data from JLE file
                                from config import extract_jle_data
                                # Create a file-like object for the JLE content
                                import io
                                jle_file_obj = io.BytesIO(jle_content)
                                jle_file_obj.name = jle_filename or 'extracted.jle'  # Set the original name
                                jle_data = extract_jle_data(jle_file_obj)

                                # Store JLE data in session state
                                st.session_state.jle_data = jle_data

                                # Process the files with JLE data
                                # Recreate Excel file object from stored content
                                import io
                                excel_file_bytes = st.session_state.excel_content
                                excel_file_obj = io.BytesIO(excel_file_bytes)
                                excel_file_obj.name = st.session_state.get('excel_filename', 'excel_upload.xlsx')

                                selected_dbf_file = st.session_state.selected_dbf_file

                                # Get the DBF content
                                dbf_content = selected_dbf_file.getvalue()

                                # We need to create a file-like object for the DBF content
                                dbf_file_obj = io.BytesIO(dbf_content)
                                dbf_file_obj.name = st.session_state.selected_dbf_name

                                updated_dbf_bytes, matched_count = process_files_with_jle(jle_data, excel_file_obj, dbf_file_obj, st.session_state.selected_dbf_name)

                                if matched_count > 0:
                                    st.success(f"Successfully processed! Matched and updated {matched_count} rows.")
                                else:
                                    st.warning(f"Files processed but no matches found. This might indicate that the ID values in your Excel file don't match those in your DBF file.")

                                # Use the selected DBF filename as output
                                output_filename = st.session_state.selected_dbf_name

                                # Provide download link for the updated DBF file
                                st.download_button(
                                    label="Download Updated DBF",
                                    data=updated_dbf_bytes,
                                    file_name=output_filename,
                                    mime="application/octet-stream"
                                )

                                # Show DBF viewer after update
                                with st.container(border=True):
                                    st.subheader("📋 Updated DBF Content")
                                    try:
                                        df = read_dbf_to_dataframe(updated_dbf_bytes)

                                        # Display the dataframe
                                        st.dataframe(df, use_container_width=True, height=400)

                                        # Store the dataframe in session state for later use
                                        st.session_state.df = df
                                        st.session_state.updated_dbf_bytes = updated_dbf_bytes

                                    except Exception as e:
                                        st.error(f"Could not display DBF content: {str(e)}")

                            finally:
                                # Clean up temporary files
                                try:
                                    os.unlink(jle_path)
                                    os.unlink(dbf_path)
                                except:
                                    pass  # Ignore errors during cleanup

                    except Exception as e:
                        st.error(f"Error processing files: {str(e)}")
                        # Optionally log the full traceback for debugging
                        import traceback
                        st.error(f"Full error details: {traceback.format_exc()}")

                    # After successful update, automatically generate Word report from template and convert to PDF
                    if 'df' in st.session_state and st.session_state.df is not None:
                        df = st.session_state.df

                        # Get JLE data if available
                        jle_data = st.session_state.get('jle_data', None) if 'jle_data' in st.session_state else None

                        # Generate Word report from template and convert to PDF
                        with st.spinner('Generating Word report from template and converting to PDF...'):
                            try:
                                # Check if we have JLE data to use the new matching functionality
                                if jle_data:
                                    # Use the new JLE/DBF matching functionality with uploaded files
                                    from reports import generate_word_report_from_jle_and_uploaded_dbf, find_matching_dbf_from_jle_data
                                    # Use the original DBF filename for pattern matching
                                    original_dbf_filename = st.session_state.selected_dbf_name if 'selected_dbf_name' in st.session_state else None

                                    # Debug: Show what we're trying to match
                                    if original_dbf_filename:
                                        import re
                                        parts = original_dbf_filename.replace('.DBF', '').replace('.dbf', '').split('_')
                                        if len(parts) >= 4:
                                            st.info(f"🔍 Attempting to match: Year/Sem={parts[1]}, SubjNum={parts[2]}, SubjCode={parts[3]}")

                                    # Debug: Show what courses are available in JLE data
                                    if 'course_data' in jle_data and jle_data['course_data'] is not None and not jle_data['course_data'].empty:
                                        jle_df = jle_data['course_data']
                                        st.info(f"📚 JLE contains {len(jle_df)} course(s): {[row['Subject Code'] + '(' + row['Subject Num'] + ')' for _, row in jle_df.iterrows()]}")

                                    word_bytes = generate_word_report_from_jle_and_uploaded_dbf(jle_data, original_dbf_filename, df)

                                    # Check if matching failed by looking for error indicators in the generated document
                                    import zipfile
                                    import io

                                    # Extract the document content to check for matching failure
                                    with zipfile.ZipFile(io.BytesIO(word_bytes)) as docx_zip:
                                        # Read the main document XML
                                        doc_xml = docx_zip.read('word/document.xml')
                                        doc_content = doc_xml.decode('utf-8', errors='ignore')

                                        # Check if error indicators are present (meaning matching failed)
                                        if 'NO MATCH FOUND' in doc_content:
                                            st.warning("⚠️ Warning: Could not match DBF file with JLE data. Please verify your file naming conventions follow the pattern: ORG_YYYYX_SUBJNUM_SUBJCODE_ID.DBF")
                                        else:
                                            # If matching succeeded, let's also check headers/footers
                                            try:
                                                header_footer_content = ""
                                                if 'word/header1.xml' in docx_zip.namelist():
                                                    header_content = docx_zip.read('word/header1.xml').decode('utf-8', errors='ignore')
                                                    header_footer_content += header_content
                                                if 'word/footer1.xml' in docx_zip.namelist():
                                                    footer_content = docx_zip.read('word/footer1.xml').decode('utf-8', errors='ignore')
                                                    header_footer_content += footer_content

                                                # Check if [Insert] placeholders still exist in headers/footers (meaning replacement failed there)
                                                import re
                                                remaining_placeholders = re.findall(r'\[Insert [^\]]*\]', header_footer_content)
                                                if remaining_placeholders:
                                                    st.warning(f"⚠️ Warning: Found {len(remaining_placeholders)} unfilled placeholders in headers/footers: {list(set(remaining_placeholders))[:5]}...")  # Show first 5 unique
                                                    st.info("This may indicate the template doesn't have the expected [Insert *] placeholders in headers/footers")
                                            except:
                                                pass  # If there's an issue reading headers/footers, continue anyway
                                else:
                                    # Use the original functionality
                                    word_bytes = generate_word_report(df, jle_data)

                                # Store the Word doc in session state
                                st.session_state.word_bytes = word_bytes
                                st.session_state.word_filename = f"E-Class_Report_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.docx"
                                st.session_state.word_report_generated = True

                                # Convert Word to PDF
                                pdf_from_word_bytes = convert_word_to_pdf(word_bytes)

                                # Store the PDF from Word in session state
                                st.session_state.pdf_from_word_bytes = pdf_from_word_bytes
                                st.session_state.pdf_from_word_filename = f"E-Class_Report_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                                st.session_state.pdf_from_word_generated = True

                                st.success("Word report and PDF generated successfully!")

                            except Exception as e:
                                st.error(f"Error generating Word report: {str(e)}")

        else:
            st.info("Please upload a RAR file with JLE and DBF files first.")




    # Check if we have a dataframe to work with
    if 'df' in st.session_state and st.session_state.df is not None:
        df = st.session_state.df

        # Get JLE data if available
        jle_data = st.session_state.get('jle_data', None) if 'jle_data' in st.session_state else None

        # Show PDF from Word report if it was generated
        if st.session_state.get('pdf_from_word_generated', False) and 'pdf_from_word_bytes' in st.session_state:
            # Show PDF from Word report if it was generated
            with st.container(border=True):
                st.subheader("📄 Generated PDF Report")

                # Convert PDF bytes to base64 for embedding in HTML
                import base64
                pdf_bytes = st.session_state.pdf_from_word_bytes
                if isinstance(pdf_bytes, str):
                    pdf_bytes = pdf_bytes.encode('latin-1')
                elif isinstance(pdf_bytes, bytearray):
                    pdf_bytes = bytes(pdf_bytes)

                pdf_base64 = base64.b64encode(pdf_bytes).decode('utf-8')

                # Embed the PDF in an iframe
                pdf_display = f'<iframe src="data:application/pdf;base64,{pdf_base64}" width="100%" height="600" type="application/pdf">'
                st.markdown(pdf_display, unsafe_allow_html=True)

                # Download button for PDF from Word
                st.download_button(
                    label="📥 Download PDF Report",
                    data=st.session_state.pdf_from_word_bytes,
                    file_name=st.session_state.pdf_from_word_filename,
                    mime="application/pdf",
                    key="download_pdf_from_word_btn"
                )


if __name__ == "__main__":
    main()