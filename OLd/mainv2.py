import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from dbf import Table, READ_WRITE

# --- Helper function to clean numeric values ---
def clean_value(val):
    if val is None:
        return ''
    if isinstance(val, (int, float)):
        return f"{val:.1f}"
    return val

def open_excel_file():
    filepath = filedialog.askopenfilename(
        filetypes=[("Excel Files", "*.xlsx *.xlsm *.xls")],
        title="Select E-Class Record"
    )
    if filepath:
        eclass_entry.delete(0, tk.END)
        eclass_entry.insert(0, filepath)

def open_dbf_file():
    filepath = filedialog.askopenfilename(
        filetypes=[("DBF Files", "*.dbf")],
        title="Select Grade Sheet"
    )
    if filepath:
        grade_entry.delete(0, tk.END)
        grade_entry.insert(0, filepath)

def button1_action():
    eclass_entry.delete(0, tk.END)
    grade_entry.delete(0, tk.END)

def button2_action():
    excel_path = eclass_entry.get().strip()
    dbf_path = grade_entry.get().strip()

    if not excel_path or not dbf_path:
        messagebox.showwarning("Missing Input", "Please select both files.")
        return

    try:
        # Load Excel - using try-finally to ensure it's properly closed
        wb = load_workbook(excel_path, data_only=True)
        try:
            ws = wb["FFG"]

            # Find the columns with "EG" and "REMARKS" in row 7
            col_grade_idx = None
            col_remark_idx = None

            for col_idx in range(1, ws.max_column + 1):  # Iterate through all columns
                cell_value = ws.cell(row=7, column=col_idx).value
                if cell_value and str(cell_value).strip().upper() == "EG":
                    col_grade_idx = col_idx
                elif cell_value and str(cell_value).strip().upper() == "REMARKS":
                    col_remark_idx = col_idx

            if col_grade_idx is None:
                raise ValueError("Column with 'EG' header not found in row 7")
            if col_remark_idx is None:
                raise ValueError("Column with 'REMARKS' header not found in row 7")

            # Convert column indices to letters for easier reference
            def col_to_letter(col_num):
                result = ""
                while col_num > 0:
                    col_num -= 1
                    result = chr(col_num % 26 + ord('A')) + result
                    col_num //= 26
                return result

            col_grade_letter = col_to_letter(col_grade_idx)
            col_remark_letter = col_to_letter(col_remark_idx)

            row = 11
            excel_data = {}

            while True:
                cell_val = ws[f'C{row}'].value
                if cell_val is None:
                    break

                try:
                    id_str = str(int(cell_val)).strip()
                except Exception:
                    row += 1
                    continue

                col_grade = ws[f'{col_grade_letter}{row}'].value
                col_remark = ws[f'{col_remark_letter}{row}'].value
                excel_data[id_str] = (col_grade, col_remark)
                row += 1
        finally:
            wb.close()  # Ensure workbook is closed even if an exception occurs

        # Open DBF for read/write
        table = Table(dbf_path)
        try:
            table.open(mode=READ_WRITE)

            id_index = 5  # 5th column, 0-based index
            target_col3 = 2  # 3rd column (write grade)
            target_col4 = 3  # 4th column (write remark)

            matched = 0
            for record in table:
                dbf_id = str(record[id_index]).strip()
                if dbf_id in excel_data:
                    grade_val, remark_val = excel_data[dbf_id]
                    with record:  # <- Required for safe writing
                        if grade_val is not None:
                            record[target_col3] = clean_value(grade_val)
                        if remark_val is not None:
                            record[target_col4] = clean_value(remark_val)
                    matched += 1
        finally:
            table.close()  # Ensure table is closed even if an exception occurs

        messagebox.showinfo("Success", f"Matched and updated {matched} rows.")

    except Exception as e:
        messagebox.showerror("Error", str(e))

# ----- UI Setup -----
root = tk.Tk()
root.title("E-Class DBF Updater")
root.geometry("460x140")
root.resizable(False, False)
root.grid_columnconfigure(1, weight=1)

# E-Class Record row
tk.Label(root, text="E-Class Record:").grid(row=0, column=0, padx=(10,5), pady=10, sticky="e")
eclass_entry = tk.Entry(root)
eclass_entry.grid(row=0, column=1, padx=5, sticky="we")
tk.Button(root, text="Browse", width=10, command=open_excel_file).grid(row=0, column=2, padx=(5,10))

# Grade Sheet row
tk.Label(root, text="Grade Sheet:").grid(row=1, column=0, padx=(10,5), pady=5, sticky="e")
grade_entry = tk.Entry(root)
grade_entry.grid(row=1, column=1, padx=5, sticky="we")
tk.Button(root, text="Browse", width=10, command=open_dbf_file).grid(row=1, column=2, padx=(5,10))

# Buttons row
button_frame = tk.Frame(root)
button_frame.grid(row=2, column=0, columnspan=3, pady=15)
tk.Button(button_frame, text="Clear", width=15, command=button1_action).pack(side="left", padx=10)
tk.Button(button_frame, text="Update DBF", width=15, command=button2_action).pack(side="left", padx=10)

root.mainloop()
